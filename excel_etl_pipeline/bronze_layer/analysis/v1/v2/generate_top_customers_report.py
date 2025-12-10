import pandas as pd
import numpy as np
from datetime import datetime
import os
import glob
from pathlib import Path
import openpyxl


def read_parquet_data(source_directory):
    """
    Read all parquet files from the specified directory
    """
    parquet_files = glob.glob(os.path.join(source_directory, "*.parquet"))

    if not parquet_files:
        print(f"No parquet files found in {source_directory}")
        return None

    print(f"Found {len(parquet_files)} parquet files")

    # Read and concatenate all parquet files
    data_frames = []
    for file_path in parquet_files:
        try:
            df = pd.read_parquet(file_path)
            data_frames.append(df)
            print(f"Successfully read: {os.path.basename(file_path)}")
        except Exception as e:
            print(f"Error reading {file_path}: {e}")

    if not data_frames:
        print("No data could be read from parquet files")
        return None

    # Combine all dataframes
    combined_df = pd.concat(data_frames, ignore_index=True)
    print(f"Total records loaded: {len(combined_df):,}")

    return combined_df


def standardize_column_names(df):
    """
    Standardize column names to expected format
    """
    print("\nStandardizing column names...")

    # Map known column name variations
    column_mapping = {}

    # Check for CUS_CODE variations
    cus_code_cols = [
        col
        for col in df.columns
        if any(x in col.lower() for x in ["cus", "customer", "cust"])
        and "code" in col.lower()
    ]
    if cus_code_cols:
        column_mapping[cus_code_cols[0]] = "CUS_CODE"
        print(f"  Using '{cus_code_cols[0]}' as CUS_CODE")

    # Check for CUSTOMER_NAME variations
    customer_name_cols = [
        col
        for col in df.columns
        if "name" in col.lower()
        and any(x in col.lower() for x in ["customer", "cus", "cust"])
    ]
    if customer_name_cols:
        column_mapping[customer_name_cols[0]] = "CUSTOMER_NAME"
        print(f"  Using '{customer_name_cols[0]}' as CUSTOMER_NAME")

    # Check for AMOUNT variations
    amount_cols = [
        col for col in df.columns if "amount" in col.lower() or "total" in col.lower()
    ]
    if amount_cols:
        column_mapping[amount_cols[0]] = "AMOUNT"
        print(f"  Using '{amount_cols[0]}' as AMOUNT")

    # Check for date variations
    date_cols = [
        col
        for col in df.columns
        if any(x in col.lower() for x in ["date", "trn", "transaction", "time"])
    ]
    if date_cols:
        column_mapping[date_cols[0]] = "TRN_DATE"
        print(f"  Using '{date_cols[0]}' as TRN_DATE")

    # Check for branch variations
    branch_cols = [col for col in df.columns if "branch" in col.lower()]
    if branch_cols:
        column_mapping[branch_cols[0]] = "BRANCH_NAME"
        print(f"  Using '{branch_cols[0]}' as BRANCH_NAME")

    # Rename columns
    df = df.rename(columns=column_mapping)

    return df


def process_data(df):
    """
    Process and clean the data
    """
    print("\nProcessing data...")

    # Ensure required columns exist
    required_cols = ["CUS_CODE", "CUSTOMER_NAME", "AMOUNT"]
    for col in required_cols:
        if col not in df.columns:
            df[col] = "Unknown" if col in ["CUS_CODE", "CUSTOMER_NAME"] else 0

    # Convert date if available
    if "TRN_DATE" in df.columns:
        df["TRN_DATE"] = pd.to_datetime(df["TRN_DATE"], errors="coerce")
        df["YEAR"] = df["TRN_DATE"].dt.year
        df["MONTH"] = df["TRN_DATE"].dt.month
        df["YEAR_MONTH"] = df["TRN_DATE"].dt.to_period("M")

        # Get date range
        min_date = df["TRN_DATE"].min()
        max_date = df["TRN_DATE"].max()
        print(f"  Date range: {min_date} to {max_date}")
    else:
        # If no date, use current year and assume all data is for current year
        current_year = datetime.now().year
        df["YEAR"] = current_year
        df["MONTH"] = 1  # Default to January
        df["YEAR_MONTH"] = pd.Period(f"{current_year}-01", freq="M")
        print(f"  No date column found. Using year: {current_year}")

    # Fill missing values
    df["CUS_CODE"] = df["CUS_CODE"].fillna("UNKNOWN")
    df["CUSTOMER_NAME"] = df["CUSTOMER_NAME"].fillna("Unknown Customer")
    df["AMOUNT"] = pd.to_numeric(df["AMOUNT"], errors="coerce").fillna(0)

    # Handle branch information
    if "BRANCH_NAME" not in df.columns or df["BRANCH_NAME"].isnull().all():
        df["BRANCH_NAME"] = "HQ"
        print("  No branch information found. Using 'HQ' as default branch.")

    # Clean branch names
    df["BRANCH_NAME"] = (
        df["BRANCH_NAME"].fillna("Unknown Branch").astype(str).str.strip()
    )

    # Create a clean customer name by combining code and name
    df["CUSTOMER_DISPLAY"] = (
        df["CUS_CODE"].astype(str) + " - " + df["CUSTOMER_NAME"].astype(str)
    )

    print(f"\nData Summary:")
    print(f"  - Total transactions: {len(df):,}")
    print(f"  - Total amount: KES {df['AMOUNT'].sum():,.2f}")
    print(f"  - Unique customers: {df['CUS_CODE'].nunique():,}")
    print(f"  - Unique branches: {df['BRANCH_NAME'].nunique():,}")

    return df


def create_monthly_sales_report(df, output_path):
    """
    Create Excel report with monthly sales for top 20 customers per branch
    """
    print(f"\nCreating monthly sales report...")

    # Create output directory if it doesn't exist
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Get all branches
        branches = df["BRANCH_NAME"].unique()
        print(f"  Processing {len(branches)} branch(es)")

        for branch in branches:
            print(f"    Processing branch: {branch}")

            # Filter data for this branch
            branch_data = df[df["BRANCH_NAME"] == branch].copy()

            if branch_data.empty:
                print(f"      No data for branch {branch}")
                continue

            # Get top 20 customers for this branch based on total sales
            customer_totals = (
                branch_data.groupby(["CUS_CODE", "CUSTOMER_NAME"])["AMOUNT"]
                .sum()
                .reset_index()
            )
            top_20_customers = customer_totals.nlargest(20, "AMOUNT")

            if top_20_customers.empty:
                print(f"      No customer data for branch {branch}")
                continue

            # Get all unique year-month periods in the data
            year_months = sorted(branch_data["YEAR_MONTH"].unique())

            # Create a pivot table for monthly sales
            monthly_sales = branch_data.pivot_table(
                index=["CUS_CODE", "CUSTOMER_NAME"],
                columns="YEAR_MONTH",
                values="AMOUNT",
                aggfunc="sum",
                fill_value=0,
            ).reset_index()

            # Filter to only top 20 customers
            monthly_sales = monthly_sales.merge(
                top_20_customers[["CUS_CODE"]], on="CUS_CODE", how="inner"
            )

            # Calculate total for each customer
            monthly_sales["Total"] = monthly_sales.select_dtypes(
                include=[np.number]
            ).sum(axis=1)

            # Sort by total descending
            monthly_sales = monthly_sales.sort_values("Total", ascending=False)

            # Format year-month columns to MM/YY format
            formatted_columns = {}
            for col in monthly_sales.columns:
                if isinstance(col, pd.Period):
                    # Format as MM/YY (e.g., 01/25 for January 2025)
                    formatted_col = f"{col.month:02d}/{str(col.year)[-2:]}"
                    formatted_columns[col] = formatted_col
                elif col == "YEAR_MONTH":
                    # Skip this if it exists as a column
                    continue

            # Rename columns
            monthly_sales = monthly_sales.rename(columns=formatted_columns)

            # Reorder columns: CUS_CODE, CUSTOMER_NAME, monthly columns sorted, Total
            monthly_cols = sorted([col for col in monthly_sales.columns if "/" in col])
            column_order = ["CUS_CODE", "CUSTOMER_NAME"] + monthly_cols + ["Total"]

            # Ensure all columns exist
            existing_cols = [
                col for col in column_order if col in monthly_sales.columns
            ]
            monthly_sales = monthly_sales[existing_cols]

            # Format the sheet name (Excel sheet names max 31 chars)
            sheet_name = str(branch)[:31]

            # Write to Excel
            monthly_sales.to_excel(writer, sheet_name=sheet_name, index=False)

            # Format the worksheet
            worksheet = writer.sheets[sheet_name]

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Format the Total column with currency format
            total_col_idx = None
            for idx, col in enumerate(monthly_sales.columns):
                if col == "Total":
                    total_col_idx = idx + 1  # +1 for Excel's 1-based indexing
                    break

            if total_col_idx:
                total_col_letter = openpyxl.utils.get_column_letter(total_col_idx)
                for row in range(2, len(monthly_sales) + 2):  # +2 for header row
                    cell = worksheet[f"{total_col_letter}{row}"]
                    cell.number_format = "#,##0.00"

            # Format monthly amount columns
            for col_idx, col_name in enumerate(monthly_sales.columns, 1):
                if "/" in col_name:  # Monthly columns
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    for row in range(2, len(monthly_sales) + 2):
                        cell = worksheet[f"{col_letter}{row}"]
                        cell.number_format = "#,##0.00"

            print(
                f"      Added sheet '{sheet_name}' with {len(monthly_sales)} customers"
            )

        # Create a summary sheet
        create_summary_sheet(df, writer)

    print(f"\nReport saved to: {output_path}")
    return output_path


def create_summary_sheet(df, writer):
    """
    Create a summary sheet with branch statistics
    """
    print("    Creating summary sheet...")

    # Branch summary
    branch_summary = (
        df.groupby("BRANCH_NAME")
        .agg(
            Total_Sales=("AMOUNT", "sum"),
            Transaction_Count=("AMOUNT", "count"),
            Unique_Customers=("CUS_CODE", "nunique"),
        )
        .reset_index()
    )

    branch_summary = branch_summary.sort_values("Total_Sales", ascending=False)

    # Overall summary
    overall_summary = pd.DataFrame(
        {
            "Metric": [
                "Total Sales",
                "Total Transactions",
                "Total Unique Customers",
                "Number of Branches",
                "Date Range",
            ],
            "Value": [
                f"KES {df['AMOUNT'].sum():,.2f}",
                f"{len(df):,}",
                f"{df['CUS_CODE'].nunique():,}",
                f"{df['BRANCH_NAME'].nunique():,}",
                f"{df['TRN_DATE'].min().date() if 'TRN_DATE' in df.columns else 'N/A'} to {df['TRN_DATE'].max().date() if 'TRN_DATE' in df.columns else 'N/A'}",
            ],
        }
    )

    # Write summary sheets
    overall_summary.to_excel(writer, sheet_name="Overall Summary", index=False)
    branch_summary.to_excel(writer, sheet_name="Branch Summary", index=False)

    # Format summary sheets
    for sheet_name in ["Overall Summary", "Branch Summary"]:
        if sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Format currency in Branch Summary
            if sheet_name == "Branch Summary":
                for row in range(2, len(branch_summary) + 2):
                    cell = worksheet[f"B{row}"]  # Total_Sales column
                    cell.number_format = "#,##0.00"


def main():
    """
    Main function to run the report generation
    """
    print("=" * 70)
    print("PHARMA WAREHOUSE - MONTHLY SALES REPORT GENERATOR")
    print("=" * 70)

    # Configuration
    source_dir = r"C:\pharma_warehouse\excel_etl_pipeline\bronze_layer\bronze_data\bronze\tables\bronze_sales_cash_invoices_summarized"
    output_dir = (
        r"C:\pharma_warehouse\excel_etl_pipeline\bronze_layer\bronze_data\reports"
    )

    # Create timestamp for output file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f"monthly_sales_report_{timestamp}.xlsx")

    # Step 1: Read data
    print("\n[1/4] Reading parquet files...")
    df = read_parquet_data(source_dir)

    if df is None or df.empty:
        print("No data available. Exiting.")
        return

    # Step 2: Standardize column names
    print("\n[2/4] Standardizing data...")
    df = standardize_column_names(df)

    # Step 3: Process data
    df = process_data(df)

    # Step 4: Generate report
    print("\n[3/4] Generating Excel report...")

    # Import openpyxl for formatting
    import openpyxl
    from openpyxl.utils import get_column_letter

    report_path = create_monthly_sales_report(df, output_file)

    print("\n[4/4] Report generation complete!")
    print("=" * 70)
    print(f"\nREPORT DETAILS:")
    print(f"  File: {report_path}")
    print(f"  Branches processed: {df['BRANCH_NAME'].nunique()}")
    print(f"  Total customers analyzed: {df['CUS_CODE'].nunique()}")
    print(
        f"  Time period covered: {df['YEAR_MONTH'].min()} to {df['YEAR_MONTH'].max()}"
    )

    # Display sample of top customers across all branches
    print("\nTOP CUSTOMERS ACROSS ALL BRANCHES:")
    all_customer_totals = (
        df.groupby(["CUS_CODE", "CUSTOMER_NAME"])["AMOUNT"]
        .sum()
        .nlargest(10)
        .reset_index()
    )
    for i, row in all_customer_totals.iterrows():
        print(
            f"  {i+1:2d}. {row['CUSTOMER_NAME'][:40]:40} | KES {row['AMOUNT']:12,.2f}"
        )


if __name__ == "__main__":
    main()
