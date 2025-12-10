import pandas as pd
import numpy as np
from datetime import datetime
import os
import glob
from pathlib import Path
import warnings

warnings.filterwarnings("ignore")


def read_parquet_data_with_duplicates(source_directory):
    """
    Read all parquet files from the specified directory, handling duplicate columns
    """
    parquet_files = glob.glob(os.path.join(source_directory, "*.parquet"))

    if not parquet_files:
        print(f"No parquet files found in {source_directory}")
        return None

    print(f"Found {len(parquet_files)} parquet files")

    # Read and concatenate all parquet files
    data_frames = []
    for file_path in parquet_files:
        try:
            print(f"Reading: {os.path.basename(file_path)}")

            # Read parquet file with low_memory to handle duplicate columns
            df = pd.read_parquet(file_path)

            # Check for duplicate column names
            if len(df.columns) != len(set(df.columns)):
                print(
                    f"  Warning: Duplicate column names found in {os.path.basename(file_path)}"
                )
                # Rename duplicate columns
                cols = pd.Series(df.columns)
                for dup in cols[cols.duplicated()].unique():
                    cols[cols[cols == dup].index.values.tolist()] = [
                        f"{dup}_{i}" if i != 0 else dup for i in range(sum(cols == dup))
                    ]
                df.columns = cols
                print(f"  Renamed duplicate columns")

            data_frames.append(df)
            print(
                f"  Successfully read with {len(df)} records, {len(df.columns)} columns"
            )

        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            # Try alternative reading method
            try:
                df = pd.read_parquet(file_path, engine="pyarrow")
                data_frames.append(df)
                print(f"  Successfully read with pyarrow engine")
            except Exception as e2:
                print(f"  Also failed with pyarrow: {e2}")

    if not data_frames:
        print("No data could be read from parquet files")
        return None

    # Combine all dataframes
    combined_df = pd.concat(data_frames, ignore_index=True, sort=False)
    print(f"Total records loaded: {len(combined_df):,}")
    print(f"Total columns: {len(combined_df.columns)}")

    # Display column names
    print("\nColumns in the dataset:")
    for i, col in enumerate(combined_df.columns, 1):
        print(f"  {i:3d}. {col}")

    return combined_df


def find_and_map_columns(df):
    """
    Find and map relevant columns from the dataframe
    """
    print("\nSearching for relevant columns...")

    # Create a copy to avoid modifying original
    df_processed = df.copy()

    # Find potential columns for each required field
    column_mapping = {}

    # 1. Find CUS_CODE
    cus_code_patterns = [
        "CUS_CODE",
        "CUSTOMER_CODE",
        "CUSCODE",
        "CUSTOMERCODE",
        "CODE",
        "ID",
    ]
    for pattern in cus_code_patterns:
        matches = [col for col in df.columns if pattern.lower() in col.lower()]
        if matches:
            column_mapping["CUS_CODE"] = matches[0]
            print(f"  Found CUS_CODE: {matches[0]}")
            break

    # 2. Find CUSTOMER_NAME
    customer_name_patterns = ["CUSTOMER_NAME", "CUS_NAME", "NAME", "CUSTOMER", "CLIENT"]
    for pattern in customer_name_patterns:
        matches = [col for col in df.columns if pattern.lower() in col.lower()]
        if matches:
            column_mapping["CUSTOMER_NAME"] = matches[0]
            print(f"  Found CUSTOMER_NAME: {matches[0]}")
            break

    # 3. Find AMOUNT
    amount_patterns = ["AMOUNT", "TOTAL", "SUM", "VALUE", "PRICE", "COST"]
    for pattern in amount_patterns:
        matches = [col for col in df.columns if pattern.lower() in col.lower()]
        if matches:
            column_mapping["AMOUNT"] = matches[0]
            print(f"  Found AMOUNT: {matches[0]}")
            break

    # 4. Find DATE
    date_patterns = [
        "DATE",
        "TRN_DATE",
        "TRANSACTION_DATE",
        "TIME",
        "CREATED",
        "TIMESTAMP",
    ]
    for pattern in date_patterns:
        matches = [col for col in df.columns if pattern.lower() in col.lower()]
        if matches:
            column_mapping["TRN_DATE"] = matches[0]
            print(f"  Found TRN_DATE: {matches[0]}")
            break

    # 5. Find BRANCH
    branch_patterns = ["BRANCH", "BRANCH_NAME", "LOCATION", "STORE", "OUTLET"]
    for pattern in branch_patterns:
        matches = [col for col in df.columns if pattern.lower() in col.lower()]
        if matches:
            column_mapping["BRANCH_NAME"] = matches[0]
            print(f"  Found BRANCH_NAME: {matches[0]}")
            break

    # Apply mapping
    for new_name, old_name in column_mapping.items():
        df_processed[new_name] = df_processed[old_name]

    return df_processed, column_mapping


def process_data(df):
    """
    Process and clean the data
    """
    print("\nProcessing data...")

    # Find and map columns
    df_processed, mapping = find_and_map_columns(df)

    # Check which columns were found
    missing_cols = []
    for col in ["CUS_CODE", "CUSTOMER_NAME", "AMOUNT", "TRN_DATE", "BRANCH_NAME"]:
        if col not in df_processed.columns:
            missing_cols.append(col)
            # Add placeholder
            if col in ["CUS_CODE", "CUSTOMER_NAME"]:
                df_processed[col] = f"UNKNOWN_{col}"
            elif col == "AMOUNT":
                df_processed[col] = 0
            elif col == "BRANCH_NAME":
                df_processed[col] = "HQ"

    if missing_cols:
        print(f"  Warning: Could not find columns: {missing_cols}")
        print(f"  Using placeholders for missing columns")

    # Convert date if available
    if "TRN_DATE" in df_processed.columns and df_processed["TRN_DATE"].notna().any():
        try:
            df_processed["TRN_DATE"] = pd.to_datetime(
                df_processed["TRN_DATE"], errors="coerce"
            )
            print(f"  Successfully parsed dates")
        except Exception as e:
            print(f"  Error parsing dates: {e}")
            df_processed["TRN_DATE"] = pd.NaT

    # If no valid dates, use current year
    if "TRN_DATE" not in df_processed.columns or df_processed["TRN_DATE"].isna().all():
        current_date = datetime.now()
        df_processed["TRN_DATE"] = current_date
        print(f"  No valid dates found. Using current date: {current_date}")

    # Extract year and month
    df_processed["YEAR"] = df_processed["TRN_DATE"].dt.year
    df_processed["MONTH"] = df_processed["TRN_DATE"].dt.month
    df_processed["YEAR_MONTH"] = df_processed["TRN_DATE"].dt.to_period("M")

    # Get date range
    min_date = df_processed["TRN_DATE"].min()
    max_date = df_processed["TRN_DATE"].max()
    print(f"  Date range: {min_date} to {max_date}")

    # Convert AMOUNT to numeric
    df_processed["AMOUNT"] = pd.to_numeric(df_processed["AMOUNT"], errors="coerce")

    # Fill missing values
    df_processed["CUS_CODE"] = df_processed["CUS_CODE"].fillna("UNKNOWN").astype(str)
    df_processed["CUSTOMER_NAME"] = (
        df_processed["CUSTOMER_NAME"].fillna("Unknown Customer").astype(str)
    )
    df_processed["AMOUNT"] = df_processed["AMOUNT"].fillna(0)
    df_processed["BRANCH_NAME"] = (
        df_processed["BRANCH_NAME"].fillna("HQ").astype(str).str.strip()
    )

    # Clean customer names and codes
    df_processed["CUS_CODE"] = df_processed["CUS_CODE"].str.strip()
    df_processed["CUSTOMER_NAME"] = df_processed["CUSTOMER_NAME"].str.strip()

    print(f"\nData Summary:")
    print(f"  - Total transactions: {len(df_processed):,}")
    print(f"  - Total amount: KES {df_processed['AMOUNT'].sum():,.2f}")
    print(f"  - Unique customers: {df_processed['CUS_CODE'].nunique():,}")
    print(f"  - Unique branches: {df_processed['BRANCH_NAME'].nunique():,}")
    print(
        f"  - Date range: {df_processed['YEAR_MONTH'].min()} to {df_processed['YEAR_MONTH'].max()}"
    )

    # Show sample of data
    print(f"\nSample of processed data:")
    print(
        df_processed[["CUS_CODE", "CUSTOMER_NAME", "AMOUNT", "TRN_DATE", "BRANCH_NAME"]]
        .head(3)
        .to_string()
    )

    return df_processed


def create_monthly_sales_report(df, output_path):
    """
    Create Excel report with monthly sales for top 30 customers per branch
    """
    print(f"\nCreating monthly sales report...")

    # Create output directory if it doesn't exist
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  Installing openpyxl for Excel formatting...")
        os.system("pip install openpyxl")
        import openpyxl
        from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Get all branches
        branches = df["BRANCH_NAME"].unique()
        print(f"  Processing {len(branches)} branch(es)")

        for branch in branches:
            print(f"    Processing branch: {branch}")

            # Filter data for this branch
            branch_data = df[df["BRANCH_NAME"] == branch].copy()

            if branch_data.empty:
                print(f"      No data for branch {branch}")
                continue

            # Get top 30 customers for this branch based on total sales
            customer_totals = (
                branch_data.groupby(["CUS_CODE", "CUSTOMER_NAME"])["AMOUNT"]
                .sum()
                .reset_index()
            )

            if len(customer_totals) == 0:
                print(f"      No customer data for branch {branch}")
                continue

            # Get top 30, or all if less than 30
            n_customers = min(30, len(customer_totals))
            top_customers = customer_totals.nlargest(n_customers, "AMOUNT")

            # Get all unique year-month periods in the data
            year_months = sorted(branch_data["YEAR_MONTH"].unique())

            if len(year_months) == 0:
                print(f"      No date data for branch {branch}")
                continue

            # Create a pivot table for monthly sales
            try:
                monthly_sales = branch_data.pivot_table(
                    index=["CUS_CODE", "CUSTOMER_NAME"],
                    columns="YEAR_MONTH",
                    values="AMOUNT",
                    aggfunc="sum",
                    fill_value=0,
                ).reset_index()
            except Exception as e:
                print(f"      Error creating pivot table: {e}")
                continue

            # Filter to only top customers
            monthly_sales = monthly_sales.merge(
                top_customers[["CUS_CODE"]], on="CUS_CODE", how="inner"
            )

            # Calculate total for each customer
            monthly_sales["Total"] = monthly_sales.select_dtypes(
                include=[np.number]
            ).sum(axis=1)

            # Sort by total descending
            monthly_sales = monthly_sales.sort_values("Total", ascending=False)

            # Format year-month columns to MM/YY format
            formatted_columns = {}
            for col in monthly_sales.columns:
                if isinstance(col, pd.Period):
                    # Format as MM/YY (e.g., 01/25 for January 2025)
                    formatted_col = f"{col.month:02d}/{str(col.year)[-2:]}"
                    formatted_columns[col] = formatted_col

            # Rename columns
            monthly_sales = monthly_sales.rename(columns=formatted_columns)

            # Reorder columns: CUS_CODE, CUSTOMER_NAME, monthly columns sorted, Total
            monthly_cols = sorted([col for col in monthly_sales.columns if "/" in col])
            column_order = ["CUS_CODE", "CUSTOMER_NAME"] + monthly_cols + ["Total"]

            # Ensure all columns exist
            existing_cols = [
                col for col in column_order if col in monthly_sales.columns
            ]
            monthly_sales = monthly_sales[existing_cols]

            # Format the sheet name (Excel sheet names max 31 chars)
            sheet_name = str(branch)[:31]
            # Replace invalid characters for sheet names
            invalid_chars = ["\\", "/", "*", "?", ":", "[", "]"]
            for char in invalid_chars:
                sheet_name = sheet_name.replace(char, "_")

            # Write to Excel
            try:
                monthly_sales.to_excel(writer, sheet_name=sheet_name, index=False)

                # Format the worksheet
                worksheet = writer.sheets[sheet_name]

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if (
                                cell.value is not None
                                and len(str(cell.value)) > max_length
                            ):
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Format numeric columns
                for col_idx, col_name in enumerate(monthly_sales.columns, 1):
                    if col_name in ["Total"] or "/" in col_name:
                        col_letter = get_column_letter(col_idx)
                        # Skip header row
                        for row in range(2, len(monthly_sales) + 2):
                            cell = worksheet[f"{col_letter}{row}"]
                            cell.number_format = "#,##0.00"

                print(
                    f"      Added sheet '{sheet_name}' with {len(monthly_sales)} customers"
                )

            except Exception as e:
                print(f"      Error writing sheet for branch {branch}: {e}")

        # Create a summary sheet
        create_summary_sheet(df, writer)

    print(f"\nReport saved to: {output_path}")
    return output_path


def create_summary_sheet(df, writer):
    """
    Create a summary sheet with branch statistics
    """
    print("    Creating summary sheet...")

    try:
        # Branch summary
        branch_summary = (
            df.groupby("BRANCH_NAME")
            .agg(
                Total_Sales=("AMOUNT", "sum"),
                Transaction_Count=("AMOUNT", "count"),
                Unique_Customers=("CUS_CODE", "nunique"),
            )
            .reset_index()
        )

        branch_summary = branch_summary.sort_values("Total_Sales", ascending=False)

        # Overall summary
        overall_summary = pd.DataFrame(
            {
                "Metric": [
                    "Total Sales",
                    "Total Transactions",
                    "Total Unique Customers",
                    "Number of Branches",
                    "Date Range",
                ],
                "Value": [
                    f"KES {df['AMOUNT'].sum():,.2f}",
                    f"{len(df):,}",
                    f"{df['CUS_CODE'].nunique():,}",
                    f"{df['BRANCH_NAME'].nunique():,}",
                    f"{df['TRN_DATE'].min().date()} to {df['TRN_DATE'].max().date()}",
                ],
            }
        )

        # Write summary sheets
        overall_summary.to_excel(writer, sheet_name="Overall Summary", index=False)
        branch_summary.to_excel(writer, sheet_name="Branch Summary", index=False)

        # Format summary sheets
        for sheet_name in ["Overall Summary", "Branch Summary"]:
            if sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if (
                                cell.value is not None
                                and len(str(cell.value)) > max_length
                            ):
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

    except Exception as e:
        print(f"      Error creating summary sheets: {e}")


def main():
    """
    Main function to run the report generation
    """
    print("=" * 70)
    print("PHARMA WAREHOUSE - MONTHLY SALES REPORT GENERATOR")
    print("=" * 70)

    # Configuration
    source_dir = r"C:\pharma_warehouse\excel_etl_pipeline\bronze_layer\bronze_data\bronze\tables\bronze_sales_cash_invoices_summarized"
    output_dir = (
        r"C:\pharma_warehouse\excel_etl_pipeline\bronze_layer\bronze_data\reports"
    )

    # Create timestamp for output file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f"monthly_sales_report_{timestamp}.xlsx")

    # Step 1: Read data
    print("\n[1/4] Reading parquet files...")
    df = read_parquet_data_with_duplicates(source_dir)

    if df is None or df.empty:
        print("No data available. Exiting.")
        return

    # Step 2: Process data
    print("\n[2/4] Processing data...")
    df_processed = process_data(df)

    # Step 3: Generate report
    print("\n[3/4] Generating Excel report...")
    report_path = create_monthly_sales_report(df_processed, output_file)

    print("\n[4/4] Report generation complete!")
    print("=" * 70)
    print(f"\nREPORT DETAILS:")
    print(f"  File: {report_path}")
    print(f"  Total customers analyzed: {df_processed['CUS_CODE'].nunique()}")
    print(f"  Total branches: {df_processed['BRANCH_NAME'].nunique()}")
    print(
        f"  Time period covered: {df_processed['YEAR_MONTH'].min()} to {df_processed['YEAR_MONTH'].max()}"
    )

    # Display top customers
    if "AMOUNT" in df_processed.columns and "CUSTOMER_NAME" in df_processed.columns:
        print("\nTOP 10 CUSTOMERS ACROSS ALL BRANCHES:")
        all_customer_totals = (
            df_processed.groupby(["CUS_CODE", "CUSTOMER_NAME"])["AMOUNT"]
            .sum()
            .nlargest(10)
            .reset_index()
        )
        for i, row in all_customer_totals.iterrows():
            customer_name = str(row["CUSTOMER_NAME"])[:40]
            print(f"  {i+1:2d}. {customer_name:40} | KES {row['AMOUNT']:12,.2f}")


if __name__ == "__main__":
    main()
